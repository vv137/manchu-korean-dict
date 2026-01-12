#!/usr/bin/env python3
"""
만주어 사전 JSON ↔ Excel 변환기

사용법:
    uv run dict_converter.py json2excel <input.json> <output.xlsx> [--lang ko|en]
    uv run dict_converter.py excel2json <input.xlsx> <output.json> [--lang ko|en]
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

# 언어별 설정
LANG_CONFIG = {
    "ko": {
        "meaning_col": "meaning_ko",
        "translation_col": "translation_ko",
    },
    "en": {
        "meaning_col": "meaning_en",
        "translation_col": "translation_en",
    },
}


def json_to_excel(json_path: Path, excel_path: Path, lang: str) -> None:
    """JSON 사전 파일을 Excel로 변환합니다."""
    config = LANG_CONFIG[lang]

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # 시트 1: entries (표제어 및 의미)
    ws_entries = wb.active
    ws_entries.title = "entries"
    ws_entries.append(["id", "headword", "romanization", "meaning_num", config["meaning_col"]])

    # 시트 2: examples (예문)
    ws_examples = wb.create_sheet("examples")
    ws_examples.append(["id", "manchu", config["translation_col"]])

    for entry in data.get("entries", []):
        entry_id = entry.get("id", 0)
        headword = entry.get("headword", "")
        romanization = entry.get("romanization", "")

        # 의미 처리 (지정된 언어만 추출)
        meanings = entry.get("meanings", [])
        target_meanings = [m["text"] for m in meanings if m.get("lang") == lang]

        for i, meaning in enumerate(target_meanings, start=1):
            ws_entries.append([entry_id, headword, romanization, i, meaning])

        # 예문 처리
        translation_key = config["translation_col"]
        for example in entry.get("examples", []):
            manchu = example.get("manchu", "")
            translation = example.get(translation_key, "")
            ws_examples.append([entry_id, manchu, translation])

    # 열 너비 조정
    ws_entries.column_dimensions["A"].width = 8
    ws_entries.column_dimensions["B"].width = 25
    ws_entries.column_dimensions["C"].width = 25
    ws_entries.column_dimensions["D"].width = 12
    ws_entries.column_dimensions["E"].width = 60

    ws_examples.column_dimensions["A"].width = 8
    ws_examples.column_dimensions["B"].width = 50
    ws_examples.column_dimensions["C"].width = 60

    wb.save(excel_path)
    print(f"✓ Excel 파일 생성 완료: {excel_path} (언어: {lang})")


def excel_to_json(excel_path: Path, json_path: Path, lang: str) -> None:
    """Excel 파일을 JSON 사전으로 변환합니다."""
    wb = load_workbook(excel_path)

    # entries 시트 읽기
    ws_entries = wb["entries"]
    entries_data: dict[int, dict] = {}

    for row in ws_entries.iter_rows(min_row=2, values_only=True):
        entry_id, headword, romanization, meaning_num, meaning = row
        if entry_id is None:
            continue

        entry_id = int(entry_id)
        if entry_id not in entries_data:
            entries_data[entry_id] = {
                "id": entry_id,
                "headword": headword or "",
                "romanization": romanization or "",
                "meanings": [],
                "examples": [],
            }

        if meaning:
            entries_data[entry_id]["meanings"].append({
                "lang": lang,
                "text": str(meaning),
            })

    # examples 시트 읽기
    ws_examples = wb["examples"]
    translation_key = LANG_CONFIG[lang]["translation_col"]

    for row in ws_examples.iter_rows(min_row=2, values_only=True):
        entry_id, manchu, translation = row
        if entry_id is None:
            continue

        entry_id = int(entry_id)
        if entry_id not in entries_data:
            continue

        if manchu:
            entries_data[entry_id]["examples"].append({
                "manchu": str(manchu),
                translation_key: str(translation) if translation else "",
            })

    # examples가 비어있으면 키 제거
    for entry in entries_data.values():
        if not entry["examples"]:
            del entry["examples"]

    # id 순으로 정렬
    sorted_entries = sorted(entries_data.values(), key=lambda x: x["id"])

    # JSON 출력
    output = {
        "meta": {
            "id": f"manchu-{lang}",
            "name": f"만주어-{'한국어' if lang == 'ko' else '영어'} 사전",
            "name_en": f"Manchu-{'Korean' if lang == 'ko' else 'English'} Dictionary",
            "description": f"만주어-{'한국어' if lang == 'ko' else '영어'} 사전 데이터",
            "languages": ["mnc", lang],
            "version": "1.0.0",
        },
        "entries": sorted_entries,
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"✓ JSON 파일 생성 완료: {json_path} (언어: {lang})")


def main():
    parser = argparse.ArgumentParser(
        description="만주어 사전 JSON ↔ Excel 변환기"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    # json2excel 명령
    p_j2e = subparsers.add_parser("json2excel", help="JSON → Excel 변환")
    p_j2e.add_argument("input", type=Path, help="입력 JSON 파일")
    p_j2e.add_argument("output", type=Path, help="출력 Excel 파일")
    p_j2e.add_argument("--lang", choices=["ko", "en"], default="ko", help="목적 언어 (기본값: ko)")

    # excel2json 명령
    p_e2j = subparsers.add_parser("excel2json", help="Excel → JSON 변환")
    p_e2j.add_argument("input", type=Path, help="입력 Excel 파일")
    p_e2j.add_argument("output", type=Path, help="출력 JSON 파일")
    p_e2j.add_argument("--lang", choices=["ko", "en"], default="ko", help="목적 언어 (기본값: ko)")

    args = parser.parse_args()

    if args.command == "json2excel":
        json_to_excel(args.input, args.output, args.lang)
    elif args.command == "excel2json":
        excel_to_json(args.input, args.output, args.lang)


if __name__ == "__main__":
    main()
